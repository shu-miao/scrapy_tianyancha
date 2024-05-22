# -*- coding: utf-8 -*-
import re
import urllib.error
import urllib.request
import requests
import xlwt
from bs4 import BeautifulSoup
import lxml
import csv
import pandas
import numpy
from openpyxl import load_workbook
import time
import random
from fake_useragent import UserAgent
from collections import namedtuple
from lxml import etree

# 从ip池取ip

base_dict = {"北京市": "bj", "天津市": "tj", "河北省": "heb", "山西省": "sx", "内蒙古自治区": "nmg", "辽宁省": "ln",
             "吉林省": "jl",
             "黑龙江省": "hlj", "上海市": "sh", "江苏省": "js", "浙江省": "zj", "安徽省": "ah", "福建省": "fj",
             "江西省": "jx",
             "山东省": "sd", "河南省": "hen", "湖北省": "hub", "湖南省": "hun", "广东省": "gd", "广西壮族自治区": "gx",
             "海南省": "han",
             "重庆市": "cq", "四川省": "sc", "贵州省": "gz", "云南省": "yn", "西藏自治区": "xz", "陕西省": "snx",
             "甘肃省": "gs",
             "青海省": "qh", "宁夏回族自治区": "nx", "新疆维吾尔自治区": "xj", "香港特别行政区": "hk",
             "澳门特别行政区": "mo",
             "台湾省": "tw"}
cacheCode_dict = {"北京市": "00110000V2020", "天津市": "00120000V2020", "河北省": "00130000V2020",
                  "山西省": "00140000V2020", "内蒙古自治区": "00150000V2020", "辽宁省": "00210000V2020",
                  "吉林省": "00220000V2020", "黑龙江省": "00230000V2020", "上海市": "00310000V2020",
                  "江苏省": "00320000V2020", "浙江省": "00330000V2020", "安徽省": "00340000V2020",
                  "福建省": "00350000V2020",
                  "江西省": "00360000V2020", "山东省": "00370000V2020", "河南省": "00410000V2020",
                  "湖北省": "00420000V2020", "湖南省": "00430000V2020", "广东省": "00440000V2020",
                  "广西壮族自治区": "00450000V2020", "海南省": "00460000V2020", "重庆市": "00500000V2020",
                  "四川省": "00510000V2020", "贵州省": "00520000V2020", "云南省": "00530000V2020",
                  "西藏自治区": "00540000V2020", "陕西省": "00610000V2020", "甘肃省": "00620000V2020",
                  "青海省": "00630000V2020", "宁夏回族自治区": "00640000V2020",
                  "新疆维吾尔自治区": "00650000V2020", "香港特别行政区": "00810000V2020",
                  "澳门特别行政区": "00820000V2020", "台湾省": "00710000V2020"}


def get_proxy():
    '''
    需先运行 radis和 ProxyPool run.py
    :return:随机ip
    '''
    PROXY_POOL_URL = 'http://localhost:5555/random'
    try:
        response = requests.get(PROXY_POOL_URL)
        if response.status_code == 200:
            return response.text
    except ConnectionError:
        return None


def get_headers():
    ua = UserAgent()
    headers = {
        'User-Agent': ua.random,
        'cookies': 'TYCID=2993b5700eaa11efb029a3ac2bd34ac2; CUID=6291b802364d23e0355ddc2a7f1f0091; '
                   'jsid=SEO-BING-ALL-SY-000001; ssuid=1224476340; '
                   'sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%22304367727%22%2C%22first_id%22%3A'
                   '%2218f61b125bdc55-0a842dcac36385-26001d51-2073600-18f61b125bec70%22%2C%22props%22%3A%7B%22'
                   '%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22'
                   '%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93'
                   '%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%7D%2C%22identities%22%3A'
                   '%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMThmNjFiMTI1YmRjNTUtMGE4NDJkY2FjMzYzODUtMjYwMDFkNTEtMjA3MzYwMC0xOGY2MWIxMjViZWM3MCIsIiRpZGVudGl0eV9sb2dpbl9pZCI6IjMwNDM2NzcyNyJ9%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%24identity_login_id%22%2C%22value%22%3A%22304367727%22%7D%2C%22%24device_id%22%3A%2218f61b125bdc55-0a842dcac36385-26001d51-2073600-18f61b125bec70%22%7D; bannerFlag=true; HWWAFSESID=4e60834f1ade63ab0cc; HWWAFSESTIME=1715571691895; csrfToken=jEVISMf3RAXspbzXkwYFfUwk; tyc-user-phone=%255B%252218788769290%2522%255D; tyc-user-info=%7B%22state%22%3A%220%22%2C%22vipManager%22%3A%220%22%2C%22mobile%22%3A%2218788769290%22%2C%22userId%22%3A%22304367727%22%7D; tyc-user-info-save-time=1715587220913; auth_token=eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIxODc4ODc2OTI5MCIsImlhdCI6MTcxNTU4NzIyMCwiZXhwIjoxNzE4MTc5MjIwfQ.xnLAPTxcAHE8RwEcX-BUlDYbMYPKo1knh_mFNGdvYe_hUtK-3oo0XPljN12pRv1Q95m7kAaxU3QliwqDqWAMYA; searchSessionId=1715586490.34441803'
    }
    return headers


def get_ip():
    proxy = get_proxy()
    if proxy == None:
        get_ip()
    else:
        return proxy


def get_url_m(headers, ip, key, province, page):
    '''
    :param headers: 请求头
    :param ip: ip地址
    :param key: 关键词
    :param province: 省份地区
    :param page: 页数
    :return: 携带关键词和省份地区的url
    '''
    proxies = {
        'http': f'http//{ip}'
    }
    base = base_dict[province]
    cacheCode = cacheCode_dict[province]
    url = f"https://www.tianyancha.com/search?key={key}&base={base}&cacheCode={cacheCode}&pageNum={page}"
    print(f'携带ip：{ip} 访问 {url}')
    strhtml = requests.get(url, headers=headers, proxies=proxies)
    soup = BeautifulSoup(strhtml.text, 'lxml')
    # print(soup)
    info = soup.select('.index_search-item-center__Q2ai5')
    # print(info)
    A, B = 'href="', '" tar'
    url_l = re.findall(f"{A}.+?{B}", str(info))
    url_m = []
    for i in range(len(url_l)):
        j = re.sub('href="', '', str(url_l[i]))
        j = re.sub('" tar', '', j)
        url_m.append(j)
    # url_m = re.sub((f"{A}|{B}", "", str(url_l)))
    return url_m


import time
# 导入selenium包
from selenium import webdriver
from selenium.webdriver.common.by import By
import xlwt
import xlrd
import os


def get_cookie():
    browser = webdriver.Chrome()
    browser.maximize_window()
    browser.get('https://www.tianyancha.com/login')
    time.sleep(4)
    browser.implicitly_wait(5)
    workbook = xlwt.Workbook(encoding='utf-8')
    sheet = workbook.add_sheet('Sheet01')
    sheet.write(0, 0, 'number')  # 通过行列坐标写入值
    sheet.write(0, 1, 'name')
    sheet.write(0, 2, 'value')
    sheet.write(0, 3, 'path')
    sheet.write(0, 4, 'domain')
    browser.find_element(By.XPATH, '//*[@id="web-content"]/div/div/div/div/div[2]').click()
    browser.find_element(By.XPATH,
                         '//*[@id="web-content"]/div/div/div/div/div[6]/div/div[1]/div[2]').click()
    browser.find_element(By.XPATH, '//*[@id="mobile"]').send_keys('登录账号')
    browser.find_element(By.XPATH, '//*[@id="password"]').send_keys('密码')
    browser.find_element(By.XPATH, '//*[@id="agreement-checkbox-account"]').click()
    time.sleep(2)
    browser.find_element(By.XPATH,
                         '//*[@id="web-content"]/div/div/div/div/div[6]/div/div[2]/div[4]').click()
    login_elements1 = browser.find_elements(By.XPATH, "/html/body/div[9]/div[2]/div/div[1]/div[1]/div[1]")
    if login_elements1 != None:
        print("出现登录验证，请20秒内完成验证")
    time.sleep(20)
    cookies = browser.get_cookies()
    for i in range(1, len(cookies) + 1):  # 遍历cookie的值，并通过行列坐标写入值
        sheet.write(i, 0, i)
        sheet.write(i, 1, cookies[i - 1]['name'])
        sheet.write(i, 2, cookies[i - 1]['value'])
        sheet.write(i, 3, cookies[i - 1]['path'])
        sheet.write(i, 4, cookies[i - 1]['domain'])
    workbook.save('cookies.xls')


def login_findurl_m(key, province, page):
    '''
    :param key:关键词
    :param province:省份地区
    :param page: 页数
    :return: 公司详情页地址（列表）
    '''
    driver = webdriver.Chrome()  # Firefox,Ie等
    driver.maximize_window()
    base = base_dict[province]
    cacheCode = cacheCode_dict[province]
    url = f"https://www.tianyancha.com/search?key={key}&base={base}&cacheCode={cacheCode}&pageNum={page}"
    driver.get(url)
    driver.implicitly_wait(6)
    # login_elements1 = driver.find_elements(By.XPATH, r'//*[@id="captcha"]/div[3]/div[1]/div[2]')
    # if login_elements1 != None:
    #     print("出现人机验证，请20秒内完成验证")
    #     time.sleep(20)
    driver.implicitly_wait(6)
    workbook = xlrd.open_workbook('cookies.xls')  # 打开指定的excel文件
    sheet = workbook.sheet_by_name('Sheet01')  # 找到指定的sheet页
    # 遍历sheet页中有效的行，在把excel中cookie信息添加到cookie中，实现免登录
    for i in range(1, sheet.nrows):
        driver.add_cookie({'name': sheet.cell_value(i, 1), 'value': sheet.cell_value(i, 2),
                           'path': sheet.cell_value(i, 3), 'domain': sheet.cell_value(i, 4)})
    time.sleep(1)
    driver.refresh()
    driver.get(url)
    driver.implicitly_wait(6)
    # login_elements1 = driver.find_elements(By.XPATH, r'//*[@id="captcha"]/div[3]/div[1]/div[2]')
    # if login_elements1 != None:
    #     print("出现人机验证，请20秒内完成验证")
    #     time.sleep(20)
    pageSource = driver.page_source
    soup = BeautifulSoup(pageSource, 'html.parser')
    soup.prettify()
    info = soup.select('.index_name__qEdWi')
    # print(info)
    A, B = 'href="', '" tar'
    url_l = re.findall(f"{A}.+?{B}", str(info))
    url_m = []
    for i in range(len(url_l)):
        j = re.sub('href="', '', str(url_l[i]))
        j = re.sub('" tar', '', j)
        url_m.append(j)
    return url_m

    # driver.find_element(By.XPATH, '//*[@id="page-container"]/div[1]/div/div[3]/div[2]/div[1]/div/input').send_keys(
    #     '集成电路设计')
    # driver.find_element(By.XPATH, '//*[@id="page-container"]/div[1]/div/div[3]/div[2]/div[1]/button').click()
    # driver.find_element(By.XPATH,
    #                     '//*[@id="page-container"]/div/div[2]/section/main/div[1]/div/div/div[3]/div/span').click()
    # driver.find_element(By.XPATH,
    #                     '//*[@id="page-container"]/div/div[2]/section/main/div[1]/div/div/div[2]/div/div[1]/div/div['
    #                     '1]/div/div[1]/span/span').click()
    # time.sleep(20)


def login_findtext(url_m, key):
    '''
    :param url_m:公司详情网页
    :return: 公司信息列表
    '''
    driver = webdriver.Chrome()  # Firefox,Ie等
    driver.maximize_window()
    driver.get(url_m)
    time.sleep(1)
    driver.implicitly_wait(10)
    # login_elements1 = driver.find_elements(By.ID, "captcha-text")
    # if login_elements1 != None:
    #     print("出现人机验证，请20秒内完成验证")
    # time.sleep(20)
    workbook = xlrd.open_workbook('cookies.xls')  # 打开指定的excel文件
    sheet = workbook.sheet_by_name('Sheet01')  # 找到指定的sheet页
    # 遍历sheet页中有效的行，在把excel中cookie信息添加到cookie中，实现免登录
    for i in range(1, sheet.nrows):
        driver.add_cookie({'name': sheet.cell_value(i, 1), 'value': sheet.cell_value(i, 2),
                           'path': sheet.cell_value(i, 3), 'domain': sheet.cell_value(i, 4)})
    time.sleep(1)
    driver.refresh()
    driver.get(url_m)
    driver.implicitly_wait(10)
    # time.sleep(10)
    pageSource = driver.page_source
    soup = BeautifulSoup(pageSource, 'html.parser')
    soup.prettify()
    info = soup.select('.index_tableBox__ZadJW ')
    text_list = []
    A, B = '企业名称</td><td colspan="5"><div class="index_copy-box__7b6Aq"><span class="index_copy-text__ri7W6">', '</span>'
    company = re.findall(f'{A}.+?{B}', str(info))
    company = re.sub(A, '', str(company))
    company = re.sub(B, '', str(company))
    text_list.append(company[2:-2])
    A, B = 'target="_blank">', '</a></div></div><span class="index_legal-bottom-info__bYvYZ">任职<'
    legal = re.findall(f'{A}.+?{B}', str(info))
    legal = re.sub(A, '', str(legal))
    legal = re.sub(B, '', str(legal))
    text_list.append(legal[2:-2])
    A, B = '登记状态<i class="tic tic-circle-question-o index_icon-name__Khq2a"></i></td><td class="num-opening" width="264px">', '</td>'
    mode = re.findall(f'{A}.+?{B}', str(info))
    mode = re.sub(A, '', str(mode))
    mode = re.sub(B, '', str(mode))
    text_list.append(mode[2:-2])
    A, B = '成立日期</td><td>', '</td>'
    date = re.findall(f'{A}.+?{B}', str(info))
    date = re.sub(A, '', str(date))
    date = re.sub(B, '', str(date))
    text_list.append(date[2:-2])
    A, B = '统一社会信用代码<i class="tic tic-circle-question-o index_icon-name__Khq2a"></i></td><td><div class="index_copy-box__7b6Aq"><span class="index_copy-text__ri7W6">', '</span>'
    unified_code = re.findall(f'{A}.+?{B}', str(info))
    unified_code = re.sub(A, '', str(unified_code))
    unified_code = re.sub(B, '', str(unified_code))
    text_list.append(unified_code[2:-2])
    A, B = '注册资本<i class="tic tic-circle-question-o index_icon-name__Khq2a"></i></td><td width=""><div title="', '">'
    registered_capital = re.findall(f'{A}.+?{B}', str(info))
    registered_capital = re.sub(A, '', str(registered_capital))
    registered_capital = re.sub(B, '', str(registered_capital))
    text_list.append(registered_capital[2:-2])
    A, B = '<td>实缴资本</td><td>', '</td>'
    pay_capital = re.findall(f'{A}.+?{B}', str(info))
    pay_capital = re.sub(A, '', str(pay_capital))
    pay_capital = re.sub(B, '', str(pay_capital))
    text_list.append(pay_capital[2:-2])
    A, B = '工商注册号</td><td><div class="index_copy-box__7b6Aq"><span class="index_copy-text__ri7W6">', '</span>'
    business_code = re.findall(f'{A}.+?{B}', str(info))
    business_code = re.sub(A, '', str(business_code))
    business_code = re.sub(B, '', str(business_code))
    text_list.append(business_code[2:-2])
    A, B = '纳税人识别号<i class="tic tic-circle-question-o index_icon-name__Khq2a"></i></td><td><div class="index_copy-box__7b6Aq"><span class="index_copy-text__ri7W6">', '</span>'
    taxpayer_number = re.findall(f'{A}.+?{B}', str(info))
    taxpayer_number = re.sub(A, '', str(taxpayer_number))
    taxpayer_number = re.sub(B, '', str(taxpayer_number))
    text_list.append(taxpayer_number[2:-2])
    A, B = '组织机构代码<i class="tic tic-circle-question-o index_icon-name__Khq2a"></i></td><td><div class="index_copy-box__7b6Aq"><span class="index_copy-text__ri7W6">', '</span>'
    organization_code = re.findall(f'{A}.+?{B}', str(info))
    organization_code = re.sub(A, '', str(organization_code))
    organization_code = re.sub(B, '', str(organization_code))
    text_list.append(organization_code[2:-2])
    A, B = '行业</td><td>', '</td>'
    industry = re.findall(f'{A}.+?{B}', str(info))
    industry = re.sub(A, '', str(industry))
    industry = re.sub(B, '', str(industry))
    text_list.append(industry[2:-2])
    A, B = '注册地址<i class="tic tic-circle-question-o index_icon-name__Khq2a"></i></td><td colspan="3"><div class="index_copy-box__7b6Aq"><span class="index_copy-text__ri7W6">', '</span>'
    address = re.findall(f'{A}.+?{B}', str(info))
    address = re.sub(A, '', str(address))
    address = re.sub(B, '', str(address))
    text_list.append(address[2:-2])
    A, B = '经营范围<i class="tic tic-circle-question-o index_icon-name__Khq2a"></i></td><td colspan="5"><div class="index_copy-box__7b6Aq"><span class="index_copy-text__ri7W6">', '</span>'
    business_scope = re.findall(f'{A}.+?{B}', str(info))
    business_scope = re.sub(A, '', str(business_scope))
    business_scope = re.sub(B, '', str(business_scope))
    text_list.append(business_scope[2:-2])
    text_list.append(key)
    # print(soup)
    return text_list

#创建csv文件
def create_csv():
    fo = open('company_tabel.csv', 'w', newline='')
    header = ['关键词', '公司名称', '法人', '登记状态', '成立日期', '统一社会信用代码', '注册资本', '实缴资本',
              '工商注册号',
              '纳税人识别号', '组织机构代码', '行业', '注册地址', '经营范围']
    writer = csv.DictWriter(fo, header)
    writer.writeheader()
    fo.close()


create_csv()

def write_csv(list):
    '''
    :param list:待写入内容
    :return: none
    '''
    fo = open('company_tabel.csv', 'a', newline='')
    header = ['关键词', '公司名称', '法人', '登记状态', '成立日期', '统一社会信用代码', '注册资本', '实缴资本',
              '工商注册号',
              '纳税人识别号', '组织机构代码', '行业', '注册地址', '经营范围']
    writer = csv.DictWriter(fo, header)
    # writer.writeheader()
    writer.writerow({'关键词': str(list[13]), '公司名称': str(list[0]), '法人': str(list[1]), '登记状态': str(list[2]),
                     '成立日期': str(list[3]),
                     '统一社会信用代码': str(list[4]), '注册资本': str(list[5]), '实缴资本': str(list[6]),
                     '工商注册号': str(list[7]), '纳税人识别号': str(list[8]), '组织机构代码': str(list[9]),
                     '行业': str(list[10]), '注册地址': str(list[11]), '经营范围': str(list[12])})
    fo.close()


# p = login_findurl_m(key='集成电路设计', province='北京市', page='1')
# get_cookie()
def run(key_list, province_list):
    '''
    :param key_list:关键词列表
    :param province_list: 省份地区列表
    :return:None
    '''
    get_cookie()
    for i in key_list:
        for j in province_list:
            url_m1 = login_findurl_m(key=str(i), province=str(j), page='1')
            url_m2 = login_findurl_m(key=str(i), province=str(j), page='2')
            url_m1.extend(url_m2)
            for k in url_m1:
                p = login_findtext(k, i)
                write_csv(p)
                print(p)
            # get_cookie()


# key_list = ['集成电路设计']
key_list = ['集成电路设计', '晶圆制造', '集成电路封测', '第三代半导体芯片制造', '服务器主机', '服务器网卡',
            '服务器机箱', '服务器电源', '服务器电源散热器', '服务器路由器', '服务器温控设备', '智能手机', '平板电脑',
            '智能可穿戴设备', '智能机器人', '智能头盔', '大型虚拟现实设备', '大尺寸液晶电视', '裸眼3D电视',
            '4K/8K超高清显示屏', '量子点显示', '柔性OLED显示', '无屏显示', '云服务IAAS', '云服务PaaS', '云服务SaaS',
            '云安全服务', '数据资产确权', '数据评估', '数据撮合', '数据审计', '数据仲裁', '数据信托', '数据担保',
            '数据保险', '核心基础软件', '软件外包产业', '嵌入式软件', '全栈式金融国产化软件', '信创应用后台加工',
            '信创应用离线分析', '信创应用存储备份', '信创应用工业互联网', '信创应用人工智能推理', '智能制造',
            '智慧城市', '智慧医疗', '智能网联汽车', '人工智能海量训练资源库', '人工智能标准测试数据集']

provice_list = ['北京市', '天津市', '河北省', '山西省', '内蒙古自治区', '辽宁省', '吉林省', '黑龙江省', '上海市',
                '江苏省', '浙江省', '安徽省', '福建省', '江西省', '山东省', '河南省', '湖北省','湖南省', '广东省',
                '广西壮族自治区', '海南省', '重庆市', '四川省', '贵州省', '云南省', '西藏自治区', '陕西省', '甘肃省',
                '青海省', '宁夏回族自治区', '新疆维吾尔自治区']

run(key_list, provice_list)
# list1 = ['北京中关村集成电路设计园发展有限责任公司', '储鑫', '存续', '2015-02-05', '91110108330373248B', '22000万人民币', '22000万人民币', '110108018607071', '91110108330373248B', '33037324-8', '软件和信息技术服务业', '北京市海淀区丰豪东路9号院2号楼4单元608', '一般项目：技术服务、技术开发、技术咨询、技术交流、技术转让、技术推广；非居住房地产租赁；物业管理；房地产评估；房地产经纪；房地产咨询；计算机系统服务；以自有资金从事投资活动；自有资金投资的资产管理服务；园区管理服务；租赁服务（不含许可类租赁服务）；仪器仪表修理；仪器仪表销售；会议及展览服务；组织文化艺术交流活动；专业设计服务；广告制作；广告发布；广告设计、代理。（除依法须经批准的项目外，凭营业执照依法自主开展经营活动）许可项目：房地产开发经营；第一类增值电信业务；互联网信息服务；基础电信业务；第二类增值电信业务；检验检测服务。（依法须经批准的项目，经相关部门批准后方可开展经营活动，具体经营项目以相关部门批准文件或许可证件为准）（不得从事国家和本市产业政策禁止和限制类项目的经营活动。）']
# list2 = ['北京神州龙芯集成电路设计有限公司', '邢光新', '存续', '2002-08-05', '911101087415895969', '15000万人民币', '15000万人民币', '110000003998414', '911101087415895969', '74158959-6', '科技推广和应用服务业', '北京市海淀区蓝靛厂东路2号院2号楼（金源时代商务中心2号楼）5层2单元（B座）6F-1', '开发、委托生产经国家密码管理机构批准的商用密码产品；销售经国家密码管理局审批并通过指定检测机构产品质量检测的商用密码产品；技术开发、技术转让、技术咨询、技术服务；计算机系统集成；销售计算机、软件及辅助设备、电子产品、通信设备、机械设备；经济信息咨询。（知识产权出资5100万元。企业依法自主选择经营项目，开展经营活动；依法须经批准的项目，经相关部门批准后依批准的内容开展经营活动；不得从事本市产业政策禁止和限制类项目的经营活动。）']
# write_csv(list1)
# write_csv(list2)
# login_findurl_m('集成电路设计','北京市',1)

# for i in url_m1:
#     p = login_findtext(str(i))
#     print(p)

# p = login_findtext('https://www.tianyancha.com/company/6159818')
# print(p)
# browser = webdriver.Chrome()
# # 指定加载页面
# browser.maximize_window()
# browser.get("https://www.tianyancha.com/login")
# time.sleep(5)
# # browser.implicitly_wait(8)
# workbook = xlwt.Workbook(encoding='utf-8')
# sheet = workbook.add_sheet('Sheet01')
# sheet.write(0, 0, 'number')  # 通过行列坐标写入值
# sheet.write(0, 1, 'name')
# sheet.write(0, 2, 'value')
# sheet.write(0, 3, 'path')
# sheet.write(0, 4, 'domain')

# input_text = browser.find_element(By.XPATH,'//*[@id="page-container"]/div[1]/div/div[3]/div[2]/div[1]/div/input').send_keys('集成电路设计')
# find_button = browser.find_element(By.XPATH,'//*[@id="page-container"]/div[1]/div/div[3]/div[2]/div[1]/button').click()
# find_put = browser.find_element(By.XPATH,'//*[@id="page-container"]/div/div[2]/section/main/div[1]/div/div/div[3]/div/span').click()
# find_basebutton = browser.find_element(By.XPATH,'//*[@id="page-container"]/div/div[2]/section/main/div[1]/div/div/div[2]/div/div[1]/div/div[1]/div/div[1]/span/span').click()

# 通过id属性获取搜索输入框
# input_mobile = browser.find_element(By.ID, "header-company-search")
# 向搜索输入框内输入selenium
# find_button = browser.find_element(By.XPATH,'//*[@id="page-header"]/div/div[2]/div/div[6]/span').click()
# find_modebutton = browser.find_element(By.XPATH, '//*[@id="web-content"]/div/div/div/div/div[2]').click()
# find_passwordbutton = browser.find_element(By.XPATH,
#                                            '//*[@id="web-content"]/div/div/div/div/div[6]/div/div[1]/div[2]').click()
# input_mobile = browser.find_element(By.XPATH, '//*[@id="mobile"]').send_keys('18788769290')
# input_password = browser.find_element(By.XPATH, '//*[@id="password"]').send_keys('lzh20001217...')
# find_agreebutton = browser.find_element(By.XPATH, '//*[@id="agreement-checkbox-account"]').click()
# time.sleep(2)
# find_loginbutton = browser.find_element(By.XPATH,
#                                         '//*[@id="web-content"]/div/div/div/div/div[6]/div/div[2]/div[4]').click()
# login_elements1 = browser.find_elements(By.XPATH, "/html/body/div[9]/div[2]/div/div[1]/div[1]/div[1]")
# if login_elements1 != None:
#     print("出现登录验证，请20秒内完成验证")
# time.sleep(20)
# cookies = browser.get_cookies()
# for i in range(1, len(cookies) + 1):  # 遍历cookie的值，并通过行列坐标写入值
#     sheet.write(i, 0, i)
#     sheet.write(i, 1, cookies[i - 1]['name'])
#     sheet.write(i, 2, cookies[i - 1]['value'])
#     sheet.write(i, 3, cookies[i - 1]['path'])
#     sheet.write(i, 4, cookies[i - 1]['domain'])
# workbook.save('test.xls')
#
# input_text = browser.find_element(By.XPATH,
#                                   '//*[@id="page-container"]/div[1]/div/div[3]/div[2]/div[1]/div/input').send_keys(
#     '集成电路设计')
# find_button = browser.find_element(By.XPATH, '//*[@id="page-container"]/div[1]/div/div[3]/div[2]/div[1]/button').click()
# browser.implicitly_wait(5)
# find_put = browser.find_element(By.XPATH,
#                                 '//*[@id="page-container"]/div/div[2]/section/main/div[1]/div/div/div[3]/div/span').click()
# time.sleep(1)
# find_basebutton = browser.find_element(By.XPATH,
#                                        '//*[@id="page-container"]/div/div[2]/section/main/div[1]/div/div/div[2]/div/div[1]/div/div[1]/div/div[1]/span/span')
# # input_mobile.send_keys("集成电路设计")
# # find_button = browser.find_element(By.XPATH,'/html/body/div[1]/div/div[2]/div/div[2]/div[1]/div').click()
# # 设置停留五秒后执行下一步
# time.sleep(100)
# # 关闭浏览器
# browser.quit()
